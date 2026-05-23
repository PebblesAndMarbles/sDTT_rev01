import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import fnmatch

SUPPORTED_EXTENSIONS = ('.xlsx', '.xls', '.csv')

def choose_file_from_candidates(candidates, label):
    """Choose one candidate file, preferring newest when possible."""
    if not candidates:
        return None

    unique_candidates = []
    seen = set()
    for candidate in candidates:
        if candidate not in seen:
            unique_candidates.append(candidate)
            seen.add(candidate)

    if len(unique_candidates) == 1:
        return unique_candidates[0]

    unique_candidates.sort(key=lambda path: os.path.getmtime(path), reverse=True)

    print(f"Multiple {label} files found:")
    for idx, candidate in enumerate(unique_candidates, 1):
        print(f"  {idx}. {os.path.basename(candidate)}")

    while True:
        choice = input(f"Select {label} file [1-{len(unique_candidates)}] (default 1): ").strip()
        if choice == "":
            return unique_candidates[0]
        if choice.isdigit():
            choice_num = int(choice)
            if 1 <= choice_num <= len(unique_candidates):
                return unique_candidates[choice_num - 1]
        print("Invalid selection. Please try again.")

def get_directory_input():
    """Get directory path from user input"""
    while True:
        directory = input("Enter the full path to the directory containing the Excel files: ").strip()
        if os.path.exists(directory) and os.path.isdir(directory):
            return directory
        else:
            print("Invalid directory path. Please try again.")

def find_am_table_file(directory):
    """Find the old AM table file in the directory, allowing Excel or CSV."""
    pattern = re.compile(r'AM_TABLE_WW\d{2}\.(xlsx|xls|csv)$', re.IGNORECASE)
    candidates = []
    for filename in os.listdir(directory):
        if pattern.match(filename):
            candidates.append(os.path.join(directory, filename))
    return choose_file_from_candidates(candidates, "AM table")

def find_lithography_file(directory):
    """Find the lithography/new target file in the directory, allowing Excel or CSV."""
    candidates = []
    exact_names = {
        "ALL.Lithography.CD.CDSEM.MFG.CDSEM_TARGETS_1278.xlsx",
        "ALL.Lithography.CD.CDSEM.MFG.CDSEM_TARGETS_1278.xls",
        "ALL.Lithography.CD.CDSEM.MFG.CDSEM_TARGETS_1278.csv",
    }

    for filename in os.listdir(directory):
        lower_name = filename.lower()
        full_path = os.path.join(directory, filename)
        if filename in exact_names:
            candidates.append(full_path)
        elif fnmatch.fnmatch(lower_name, 'amct_target_lines_*.csv'):
            candidates.append(full_path)

    return choose_file_from_candidates(candidates, "new target")

def read_table(filepath):
    """Read CSV or Excel file into a dataframe based on file extension."""
    _, extension = os.path.splitext(filepath)
    extension = extension.lower()

    if extension == '.csv':
        return pd.read_csv(filepath)
    if extension in ('.xlsx', '.xls'):
        return pd.read_excel(filepath)

    raise ValueError(f"Unsupported file extension: {extension}")

def advanced_wildcard_match(pattern, text):
    """
    Enhanced wildcard matching that handles:
    - * : any string of characters
    - ? : single character
    - [XY] : single character that can be X or Y
    - #XXXXX : specific patterns
    """
    if pd.isna(pattern) or pd.isna(text):
        return False
    
    pattern_str = str(pattern)
    text_str = str(text)
    
    # Convert pattern to regex
    regex_pattern = ""
    i = 0
    while i < len(pattern_str):
        char = pattern_str[i]
        if char == '*':
            regex_pattern += ".*"
        elif char == '?':
            regex_pattern += "."
        elif char == '[':
            # Find the closing bracket
            j = i + 1
            while j < len(pattern_str) and pattern_str[j] != ']':
                j += 1
            if j < len(pattern_str):
                # Extract characters between brackets
                bracket_content = pattern_str[i+1:j]
                regex_pattern += f"[{re.escape(bracket_content)}]"
                i = j  # Skip to closing bracket
            else:
                regex_pattern += re.escape(char)
        elif char == '#':
            # Handle #XXXXX pattern - assuming it means exactly 5 characters
            if i + 5 < len(pattern_str) and pattern_str[i+1:i+6] == "XXXXX":
                regex_pattern += r"\d{5}"
                i += 5  # Skip the XXXXX part
            else:
                regex_pattern += re.escape(char)
        else:
            regex_pattern += re.escape(char)
        i += 1
    
    # Add anchors for exact match
    regex_pattern = f"^{regex_pattern}$"
    
    try:
        return bool(re.match(regex_pattern, text_str))
    except re.error:
        # Fallback to simple fnmatch if regex fails
        return fnmatch.fnmatch(text_str, pattern_str)

def calculate_match_score(row1, row2, columns):
    """Calculate how many columns match between two rows"""
    score = 0
    for col in columns:
        if advanced_wildcard_match(str(row1[col]), str(row2[col])):
            score += 1
    return score

def is_generic_product_row(row):
    """Check if this is a generic PRODUCT=* row"""
    return str(row['PRODUCT']).strip() == '*'

def get_pattern_specificity(pattern):
    """
    Calculate specificity of a pattern. Higher number = more specific
    * gets lowest score, exact matches get highest score
    """
    if pd.isna(pattern):
        return 0
    
    pattern_str = str(pattern).strip()
    
    if pattern_str == '*':
        return 1  # Lowest specificity
    
    specificity = 10  # Base score for non-wildcard
    
    # Count wildcards (reduce specificity)
    specificity -= pattern_str.count('*') * 3
    specificity -= pattern_str.count('?') * 1
    
    # Count specific characters (increase specificity)
    non_wildcard_chars = len(pattern_str) - pattern_str.count('*') - pattern_str.count('?')
    specificity += non_wildcard_chars
    
    return max(1, specificity)  # Ensure minimum score of 1

def filter_table1_for_specificity(table1, table2):
    """
    Filter table1 to remove generic PRODUCT=* rows when more specific matches exist
    """
    match_columns = ['PILOT_NAME', 'PRODUCT', 'LAYER', 'OPERATION', 'SPC_MEASURED_LAYER']
    
    # Group table1 rows by non-PRODUCT matching columns
    table1_groups = {}
    
    for idx1, row1 in table1.iterrows():
        # Create a key based on non-PRODUCT columns
        key_columns = [col for col in match_columns if col != 'PRODUCT']
        key = tuple(str(row1[col]) for col in key_columns)
        
        if key not in table1_groups:
            table1_groups[key] = []
        table1_groups[key].append((idx1, row1))
    
    # For each group, check if we should keep generic PRODUCT=* rows
    filtered_indices = set()
    
    for key, group_rows in table1_groups.items():
        # Check if any table2 row could match this group
        group_has_table2_matches = False
        
        for idx2, row2 in table2.iterrows():
            # Check if this table2 row matches the key columns
            key_matches = True
            key_columns = [col for col in match_columns if col != 'PRODUCT']
            
            for i, col in enumerate(key_columns):
                table1_pattern = key[i]
                if not advanced_wildcard_match(table1_pattern, str(row2[col])):
                    key_matches = False
                    break
            
            if key_matches:
                group_has_table2_matches = True
                break
        
        if not group_has_table2_matches:
            # If no table2 matches, keep all table1 rows in this group
            for idx1, row1 in group_rows:
                filtered_indices.add(idx1)
            continue
        
        # Separate generic and specific rows
        generic_rows = [(idx1, row1) for idx1, row1 in group_rows if is_generic_product_row(row1)]
        specific_rows = [(idx1, row1) for idx1, row1 in group_rows if not is_generic_product_row(row1)]
        
        # Always keep specific rows
        for idx1, row1 in specific_rows:
            filtered_indices.add(idx1)
        
        # Only keep generic rows if no specific rows exist
        if not specific_rows:
            for idx1, row1 in generic_rows:
                filtered_indices.add(idx1)
        else:
            # Check if any table2 product doesn't match any specific row
            for idx2, row2 in table2.iterrows():
                # Check if this table2 row matches the key columns
                key_matches = True
                key_columns = [col for col in match_columns if col != 'PRODUCT']
                
                for i, col in enumerate(key_columns):
                    table1_pattern = key[i]
                    if not advanced_wildcard_match(table1_pattern, str(row2[col])):
                        key_matches = False
                        break
                
                if not key_matches:
                    continue
                
                # Check if this table2 product matches any specific row
                matches_specific = False
                for idx1, row1 in specific_rows:
                    if advanced_wildcard_match(str(row1['PRODUCT']), str(row2['PRODUCT'])):
                        matches_specific = True
                        break
                
                # If no specific match found, we need the generic rows
                if not matches_specific:
                    for idx1, row1 in generic_rows:
                        filtered_indices.add(idx1)
                    break
    
    # Return filtered table1
    return table1.loc[list(filtered_indices)].reset_index(drop=True)

def find_matches_for_table2(table1, table2):
    """Find matches for each row in table2 from table1"""
    match_columns = ['PILOT_NAME', 'PRODUCT', 'LAYER', 'OPERATION', 'SPC_MEASURED_LAYER']
    matches = []
    
    # Filter table1 to remove unnecessary generic rows
    filtered_table1 = filter_table1_for_specificity(table1, table2)
    
    print(f"Filtered Table 1 from {len(table1)} to {len(filtered_table1)} rows")
    
    # For each row in table2, find the best match in filtered table1
    for idx2, row2 in table2.iterrows():
        best_match = None
        best_score = 0
        best_specificity = 0
        
        for idx1, row1 in filtered_table1.iterrows():
            score = calculate_match_score(row1, row2, match_columns)
            
            if score > 0:  # Only consider actual matches
                # Calculate specificity of the match
                specificity = sum(get_pattern_specificity(row1[col]) for col in match_columns)
                
                # Prefer higher score, then higher specificity
                if (score > best_score) or (score == best_score and specificity > best_specificity):
                    best_score = score
                    best_specificity = specificity
                    # Find original index in table1
                    original_idx = table1[
                        (table1[match_columns] == row1[match_columns]).all(axis=1)
                    ].index[0]
                    best_match = (original_idx, idx2, score)
        
        # Include the match even if score is 0 (no match found)
        if best_match:
            matches.append(best_match)
        else:
            # Create a dummy match with score 0 if no match found
            matches.append((None, idx2, 0))
    
    return matches

def create_joined_dataframes(table1, table2, matches):
    """Create joined dataframes based on matches"""
    result_table1_rows = []
    result_table2_rows = []
    match_info = []
    
    for match in matches:
        idx1, idx2, score = match
        
        # Always add the table2 row
        result_table2_rows.append(table2.iloc[idx2])
        match_info.append((idx1, idx2, score))
        
        # Add corresponding table1 row if match exists, otherwise add empty row
        if idx1 is not None:
            result_table1_rows.append(table1.iloc[idx1])
        else:
            # Create empty row with same structure as table1
            empty_row = pd.Series([None] * len(table1.columns), index=table1.columns)
            result_table1_rows.append(empty_row)
    
    result_table1 = pd.DataFrame(result_table1_rows).reset_index(drop=True)
    result_table2 = pd.DataFrame(result_table2_rows).reset_index(drop=True)
    
    return result_table1, result_table2, match_info

def create_excel_with_formatting(table1, table2, matches, output_path):
    """Create the formatted Excel file"""
    # Create joined dataframes
    joined_table1, joined_table2, match_info = create_joined_dataframes(table1, table2, matches)
    
    # Sort by PRODUCT column from table2
    sort_indices = joined_table2['PRODUCT'].argsort()
    joined_table1 = joined_table1.iloc[sort_indices].reset_index(drop=True)
    joined_table2 = joined_table2.iloc[sort_indices].reset_index(drop=True)
    sorted_match_info = [match_info[i] for i in sort_indices]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Joined_Data"
    
    # Define colors
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(color="FF0000")
    
    # Add header row
    table1_cols = len(table1.columns)
    table2_cols = len(table2.columns)
    total_cols = table1_cols + table2_cols
    
    # Merge cells for headers
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table1_cols)
    ws.merge_cells(start_row=1, start_column=table1_cols+1, end_row=1, end_column=total_cols)
    
    ws.cell(row=1, column=1, value="Old Target")
    ws.cell(row=1, column=table1_cols+1, value="New Target")
    
    # Add column headers
    for i, col in enumerate(table1.columns, 1):
        ws.cell(row=2, column=i, value=col)
    
    for i, col in enumerate(table2.columns, table1_cols+1):
        ws.cell(row=2, column=i, value=col)
    
    # Add data
    for i in range(len(joined_table1)):
        row_num = i + 3
        
        # Add table1 data
        for j, value in enumerate(joined_table1.iloc[i], 1):
            ws.cell(row=row_num, column=j, value=value)
        
        # Add table2 data
        for j, value in enumerate(joined_table2.iloc[i], table1_cols+1):
            cell = ws.cell(row=row_num, column=j, value=value)
            # Highlight TARGET column in yellow
            if table2.columns[j-table1_cols-1] == 'TARGET':
                cell.fill = yellow_fill
        
        # Highlight rows with best matches in red
        match_score = sorted_match_info[i][2]
        if match_score >= 4:  # Highlight high-scoring matches (4+ column matches)
            for j in range(1, total_cols + 1):
                ws.cell(row=row_num, column=j).font = red_font
    
    # Create deltatoNEW sheet
    delta_ws = wb.create_sheet("deltatoNEW")
    delta_ws.cell(row=1, column=1, value="Table2_PRODUCT")
    delta_ws.cell(row=1, column=2, value="Table1_PRODUCT")
    delta_ws.cell(row=1, column=3, value="TARGET_Difference")
    delta_ws.cell(row=1, column=4, value="Match_Count")
    
    # Count matches per product
    product_matches = {}
    for i, (idx1, idx2, score) in enumerate(sorted_match_info):
        product = joined_table2.iloc[i]['PRODUCT']
        if product not in product_matches:
            product_matches[product] = 0
        if idx1 is not None and score > 0:
            product_matches[product] += 1
    
    row_num = 2
    processed_products = set()
    
    for i, (idx1, idx2, score) in enumerate(sorted_match_info):
        table2_product = joined_table2.iloc[i]['PRODUCT']
        
        # Only add each product once to delta sheet
        if table2_product in processed_products:
            continue
        processed_products.add(table2_product)
        
        table1_product = joined_table1.iloc[i]['PRODUCT'] if idx1 is not None else "No Match"
        
        try:
            if idx1 is not None:
                table1_target = float(joined_table1.iloc[i]['TARGET'])
                table2_target = float(joined_table2.iloc[i]['TARGET'])
                difference = table1_target - table2_target
            else:
                difference = "No Match"
        except (ValueError, TypeError):
            difference = "N/A"
        
        match_count = product_matches.get(table2_product, 0)
        
        delta_ws.cell(row=row_num, column=1, value=table2_product)
        delta_ws.cell(row=row_num, column=2, value=table1_product)
        delta_ws.cell(row=row_num, column=3, value=difference)
        delta_ws.cell(row=row_num, column=4, value=match_count)
        row_num += 1
    
    wb.save(output_path)

def main():
    print("Excel Table Joiner with Smart Generic Filtering")
    print("===============================================")
    
    # Get directory
    directory = get_directory_input()
    
    # Find files
    am_table_file = find_am_table_file(directory)
    lithography_file = find_lithography_file(directory)
    
    if not am_table_file:
        print("AM_TABLE_WW## file not found in the selected directory.")
        return
    
    if not lithography_file:
        print("Lithography file not found in the selected directory.")
        return
    
    print(f"Found AM Table file: {os.path.basename(am_table_file)}")
    print(f"Found Lithography file: {os.path.basename(lithography_file)}")
    
    # Read Excel files
    try:
        table1 = read_table(am_table_file)
        table2 = read_table(lithography_file)
        print("Files loaded successfully.")
        print(f"Table 1 rows: {len(table1)}")
        print(f"Table 2 rows: {len(table2)}")
    except Exception as e:
        print(f"Error reading files: {e}")
        return
    
    # Find matches
    print("Finding matches between tables...")
    matches = find_matches_for_table2(table1, table2)
    print(f"Processing {len(matches)} rows from Table 2.")
    
    # Count successful matches
    successful_matches = sum(1 for match in matches if match[0] is not None and match[2] > 0)
    print(f"Found {successful_matches} successful matches.")
    
    # Create output file
    output_path = os.path.join(directory, "AM_retgt.xlsx")
    
    try:
        create_excel_with_formatting(table1, table2, matches, output_path)
        print(f"Output file created: {output_path}")
    except Exception as e:
        print(f"Error creating output file: {e}")
        return

if __name__ == "__main__":
    main()